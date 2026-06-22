#!/usr/bin/env python3
"""Generate XLSX from Procedures.xlsx template, filling M (Yes/No) and N (finding description) columns with audit data."""
import sys
import json
import os
import re
import io
from collections import OrderedDict

import psycopg2
import openpyxl
from openpyxl.styles import Font


def main():
    data = json.load(sys.stdin)
    audit_id = data["audit_id"]
    db_url = data.get("db_url", os.environ.get("DATABASE_URL", ""))
    template_path = data.get("template_path", "Procedures.xlsx")

    if not db_url:
        print("FATAL: DATABASE_URL not set", file=sys.stderr)
        sys.exit(1)

    if not os.path.exists(template_path):
        print(f"FATAL: template not found: {template_path} (cwd={os.getcwd()})", file=sys.stderr)
        sys.exit(1)

    conn = psycopg2.connect(db_url)
    cur = conn.cursor()

    cur.execute("""
        SELECT pi.id, pi.section_number, pi.sort_order, pi.control_question,
               pei.id, pei.evidence_text
        FROM procedure_items pi
        LEFT JOIN procedure_evidence_items pei ON pei.procedure_item_id = pi.id
        ORDER BY pi.sort_order, pei.sort_order
    """)

    pi_by_sort = OrderedDict()

    for row in cur.fetchall():
        pi_id, section, sort_order, question, ev_id, ev_text = row
        if sort_order not in pi_by_sort:
            pi_by_sort[sort_order] = {
                "id": pi_id,
                "section": section,
                "question": question or "",
            }

    # Fetch findings linked to controls
    cur.execute("""
        SELECT f.short_description, pi.sort_order
        FROM findings f
        JOIN procedure_items pi ON pi.id = f.procedure_item_id
        WHERE f.audit_id = %s AND f.procedure_item_id IS NOT NULL
    """, (audit_id,))

    findings_by_sort = {}
    for short_desc, sort_order in cur.fetchall():
        if sort_order not in findings_by_sort:
            findings_by_sort[sort_order] = short_desc or ""

    cur.close()
    conn.close()

    # Load template XLSX
    wb = openpyxl.load_workbook(template_path)
    ws = wb["Working Paper"]

    # Apply Tahoma 11 black to all cells
    tahoma = Font(name='Tahoma', size=11, color='000000')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if cell.value:
                cell.font = tahoma

    current_section_num = None
    control_counter = 0

    for row_idx in range(8, ws.max_row + 1):
        col_b = str(ws.cell(row=row_idx, column=2).value or "").strip()
        col_k = str(ws.cell(row=row_idx, column=11).value or "").strip()

        if not col_b and not col_k:
            continue

        sm = re.match(r"^(\d+)\.\s", col_b)
        if sm and col_k:
            current_section_num = int(sm.group(1))
            control_counter = 0
            control_counter += 1
            _fill_row(ws, row_idx, current_section_num, control_counter,
                      pi_by_sort, findings_by_sort)
        elif col_k and current_section_num:
            control_counter += 1
            _fill_row(ws, row_idx, current_section_num, control_counter,
                      pi_by_sort, findings_by_sort)

    output = io.BytesIO()
    wb.save(output)
    sys.stdout.buffer.write(output.getvalue())


def _fill_row(ws, row_idx, section_num, control_counter,
              pi_by_sort, findings_by_sort):
    sort_order = section_num * 100 + control_counter
    pi = pi_by_sort.get(sort_order)
    if not pi:
        return

    has_finding = sort_order in findings_by_sort
    m_val = "No" if has_finding else "Yes"
    n_val = findings_by_sort.get(sort_order, "")

    cell_m = ws.cell(row=row_idx, column=13)
    cell_m.value = m_val
    cell_m.font = Font(name='Tahoma', size=11, color='000000')

    cell_n = ws.cell(row=row_idx, column=14)
    cell_n.value = n_val if n_val else None
    cell_n.font = Font(name='Tahoma', size=11, color='000000')


if __name__ == "__main__":
    main()
